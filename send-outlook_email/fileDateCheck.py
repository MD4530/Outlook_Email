# send email one time a day based on file modify
# package
# templateFile =r"Q:\Manufacturing\Manheim\Shipping and Inventory\Manheim Raw Materials Inventory Transfer file - Live.xlsx"
import os
from openpyxl import load_workbook
from datetime import datetime 
import sys

def timeStampConverter(timestamp):
    return datetime.fromtimestamp(timestamp)

def fileCheck(templateFile, runTimeStoreFile):
    # time store file

    if os.path.isfile(templateFile) and os.path.isfile(runTimeStoreFile): 
        modify_time= os.path.getmtime(templateFile) #####1652690891.8609138
        
        workbook = load_workbook(runTimeStoreFile, read_only=False, data_only=True, keep_vba=True )
        sheet =workbook["Modify_Date"]
       
        maxRowNumber = sheet.max_row
        lastModifyDateInSheet =float(sheet["A" + str(maxRowNumber)].value)
        
        print ("Manheim  file update date: ", timeStampConverter( modify_time))
        print ("Last modify date in the store file: ", timeStampConverter(lastModifyDateInSheet), " Row Number: ", maxRowNumber)
        
        if lastModifyDateInSheet ==  modify_time:
            print ("File is uptodate-----------------------------------")
        
        elif lastModifyDateInSheet <  modify_time:
            print ("Need to email----------------------------")
            maxRowNumber +=1
            sheet["A" + str(maxRowNumber)] = str(modify_time)
            
            return [timeStampConverter( modify_time), workbook]
            ### return date time and workbook
            
    # ct = datetime.now().timestamp()
    